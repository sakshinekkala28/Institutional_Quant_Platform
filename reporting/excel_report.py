"""
====================================================================
Institutional Quant Platform

Excel Report

Author : Institutional Quant Platform

Purpose
-------
Institutional Excel report generator.

Exports a unified report package
to Excel format.

Inherited From

• BaseReport

====================================================================
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from reporting.base_report import BaseReport


class ExcelReport(
    BaseReport
):
    """
    Institutional Excel report.
    """

    def __init__(

        self,

        title: str,

        report_data: dict,

        author: str = "Institutional Quant Platform",

    ) -> None:

        super().__init__(

            title=title,

            author=author,

        )

        self.report_data = report_data

    # =====================================================
    # EXPORT
    # =====================================================

    def export(

        self,

        destination: str,

    ) -> None:

        path = Path(

            destination

        )

        if path.suffix.lower() != ".xlsx":

            path = path.with_suffix(

                ".xlsx"

            )

        workbook = Workbook()

        summary = workbook.active

        summary.title = "Summary"

        # =================================================
        # SUMMARY SHEET
        # =================================================

        summary["A1"] = "Title"
        summary["B1"] = self.title

        summary["A2"] = "Author"
        summary["B2"] = self.author

        summary["A3"] = "Generated"
        summary["B3"] = str(

            self.created_at

        )

        for cell in (

            summary["A1"],

            summary["A2"],

            summary["A3"],

        ):

            cell.font = Font(

                bold=True

            )

        row = 5

        summary.cell(

            row=row,

            column=1,

            value="Metadata",

        ).font = Font(

            bold=True

        )

        row += 1

        metadata = self.report_data.get(

            "Metadata",

            {},

        )

        for key, value in metadata.items():

            summary.cell(

                row=row,

                column=1,

                value=str(key),

            )

            summary.cell(

                row=row,

                column=2,

                value=str(value),

            )

            row += 1

        # =================================================
        # SECTION SHEETS
        # =================================================

        sections = self.report_data.get(

            "Sections",

            {},

        )

        for name, values in sections.items():

            sheet = workbook.create_sheet(

                title=name[:31]

            )

            sheet["A1"] = name

            sheet["A1"].font = Font(

                bold=True,

                size=14,

            )

            if isinstance(

                values,

                dict,

            ):

                sheet["A3"] = "Metric"

                sheet["B3"] = "Value"

                sheet["A3"].font = Font(

                    bold=True

                )

                sheet["B3"].font = Font(

                    bold=True

                )

                row = 4

                for key, value in values.items():

                    sheet.cell(

                        row=row,

                        column=1,

                        value=str(key),

                    )

                    sheet.cell(

                        row=row,

                        column=2,

                        value=str(value),

                    )

                    row += 1

            else:

                sheet["A3"] = str(

                    values

                )

        workbook.save(

            path

        )

    # =====================================================
    # SUMMARY
    # =====================================================

    def summary(

        self,

    ) -> dict:

        return {

            **super().summary(),

            "Sections":

                len(

                    self.report_data.get(

                        "Sections",

                        {},

                    )

                ),

            "Format":

                "Excel",

        }

    # =====================================================
    # REPRESENTATION
    # =====================================================

    def __repr__(

        self,

    ) -> str:

        return (

            f"{self.__class__.__name__}("

            f"Title='{self.title}'"

            f")"

        )

    __str__ = __repr__